import numpy as np 
import math
from scipy import integrate
#import required libraries for reading excel
from openpyxl import load_workbook
 
#read  from excel file
wb = load_workbook('data_NewEngland.xlsx')
sheet_1 = wb.get_sheet_by_name('bus')
bus = np.zeros((sheet_1.max_row,sheet_1.max_column))
 
for i in range(0,sheet_1.max_row):
	for j in range(0,sheet_1.max_column):
		bus[i,j]=sheet_1.cell(row=i+1, column=j+1).value

sheet_2 = wb.get_sheet_by_name('line')
line = np.zeros((sheet_2.max_row,sheet_2.max_column))
 
for i in range(0,sheet_2.max_row):
	for j in range(0,sheet_2.max_column):
		line[i,j]=sheet_2.cell(row=i+1, column=j+1).value

sheet_3 = wb.get_sheet_by_name('exc_con')
exc_con = np.zeros((sheet_3.max_row,sheet_3.max_column))
 
for i in range(0,sheet_3.max_row):
	for j in range(0,sheet_3.max_column):
		exc_con[i,j]=sheet_3.cell(row=i+1, column=j+1).value

sheet_4 = wb.get_sheet_by_name('ibus_con')
ibus_con = np.zeros((sheet_4.max_row,sheet_4.max_column))
 
for i in range(0,sheet_4.max_row):
	for j in range(0,sheet_4.max_column):
		ibus_con[i,j]=sheet_4.cell(row=i+1, column=j+1).value

sheet_5 = wb.get_sheet_by_name('mac_con')
mac_con = np.zeros((sheet_5.max_row,sheet_5.max_column))
 
for i in range(0,sheet_5.max_row):
	for j in range(0,sheet_5.max_column):
		mac_con[i,j]=sheet_5.cell(row=i+1, column=j+1).value

sheet_6 = wb.get_sheet_by_name('sw_con')
sw_con = np.zeros((sheet_6.max_row,sheet_6.max_column))
 
for i in range(0,sheet_6.max_row):
	for j in range(0,sheet_6.max_column):
		sw_con[i,j]=sheet_6.cell(row=i+1, column=j+1).value

sheet_7 = wb.get_sheet_by_name('bus_sol')
bus_sol = np.zeros((sheet_7.max_row,sheet_7.max_column))
 
for i in range(0,sheet_7.max_row):
	for j in range(0,sheet_7.max_column):
		bus_sol[i,j]=sheet_7.cell(row=i+1, column=j+1).value

sheet_8 = wb.get_sheet_by_name('line_flow')
line_flow = np.zeros((sheet_8.max_row,sheet_8.max_column))
 
for i in range(0,sheet_8.max_row):
	for j in range(0,sheet_8.max_column):
		line_flow[i,j]=sheet_8.cell(row=i+1, column=j+1).value

sheet_9 = wb.get_sheet_by_name('theta_orig')
theta_orig = np.zeros((sheet_9.max_row,sheet_9.max_column))
 
for i in range(0,sheet_9.max_row):
	for j in range(0,sheet_9.max_column):
		theta_orig[i,j]=sheet_9.cell(row=i+1, column=j+1).value

# finished loading data from New England

global k_control
basmva = 100
n = bus.shape[0]    
m = line.shape[0]  

G = mac_con[:,1]
G = G.astype(int)
g = G.shape[0]
L = np.arange(1,n-g+1)
l = n-g;

Vm_orig = bus_sol[:,1]
Vm_orig=Vm_orig.astype(float)
theta_orig = bus_sol[:,2]/180*math.pi

Pm_orig= bus_sol[:,3]-bus_sol[:,5]
Pm_orig=Pm_orig.astype(float)
M = np.zeros((n,n))
M[np.ix_(G-1,G-1)] = np.diag(np.multiply((2*mac_con[:,15])/(2*math.pi*60),mac_con[:,2])/basmva)
M=M.astype(float)

D=np.zeros((n,n))
D[np.ix_(G-1,G-1)] = 5.0*np.diag(bus_sol[G-1,3])/(2*math.pi*60)*np.identity(g)
D[np.ix_(L-1,L-1)] = 0.2*np.mean(np.diag(D[G-1,G-1]))*np.identity(l)
D=D*2
D=D*3
D=D.astype(float)

Tg = 0.2*np.ones((g,1))
Tb = 7.0*np.ones((g,1))
R = 1*0.05*np.divide((2*math.pi*60),bus_sol[G-1,3])
R=R.astype(float)

L_controlled=np.array([25,26])
Td=5.0*np.ones((n,1))

Bij = np.zeros((m,1))
Bij_nonlinear = np.zeros((m,1))
Bij_vf=np.zeros((m,1))
Plink_orig = np.zeros((m,1))
C= np.zeros((n,m))   
CQ= np.zeros((n,m))   
for k in range(m):
	ixp = int(line[k,0])-1
	ixn = int(line[k,1])-1
	C[ixp,k] = 1
	C[ixn,k] = -1 
	CQ[ixp,k] = 1
	CQ[ixn,k] = -1
	Bij[k]=Vm_orig[ixp]*Vm_orig[ixn]/(line[k,3]) *np.cos(theta_orig[ixp]-theta_orig[ixn])
	Bij_nonlinear[k]=Vm_orig[ixp]*Vm_orig[ixn]/(line[k,3])
	Bij_vf[k]=1.0/line[k,3]
	Plink_orig[k] = Bij_nonlinear[k]*np.sin(theta_orig[ixp]-theta_orig[ixn])

Bjj=-bus[:,8];
Bjj=Bjj.astype(float)
for cnt1 in range(n):
   for cnt2 in range(m):
   	if C[cnt1,cnt2] != 0:
   		Bjj[cnt1]= Bjj[cnt1]-Bij_vf[cnt2]

Tv = mac_con[:,8]
Tv.astype(float)
xd = mac_con[:,5]
xd = xd.astype(float)
xdt= mac_con[:,6]
xdt = xdt.astype(float)

Ef=1.0

G_controlled = np.array([ 1, 3, 5  ])
cost_alpha = bus_sol[G-1,3]
bus_controlled = mac_con[G_controlled-1,1]
bus_controlled = bus_controlled.astype(int)

limitlink=np.array([4, 19, 26])

barP = 10*np.ones((m,1))
barP[limitlink-1]= 3
underP = -barP
barDeltaP = barP - Plink_orig
underDeltaP = underP - Plink_orig

rm=np.transpose(np.arange(2,n+1))
rm=rm.astype(int)
Ct=C[rm-1,:]
needed = np.arange(1,l)
Lt=L[needed]-1

Bij = Bij.ravel()
H2 = np.matmul(Ct,np.diag(Bij))
H0 = np.linalg.inv(np.matmul(H2,np.transpose(Ct)))
H1 = np.matmul(np.transpose(Ct), H0)

H=np.transpose(np.matmul(np.diag(Bij),H1))
Hg=H[G-2,:]
Hl=H[Lt-1,:]

dtb=[30]
deltaPm=np.zeros((n,1));
deltaPm[dtb]=0; 

Tmax=50

omegao = np.zeros((g,1))
thetao = np.zeros((n,1))
thetao[:,0] = theta_orig
valveo = np.zeros((g,1))   
pmecho = np.zeros((g,1))   
loado  = np.zeros((n,1))

PF = cost_alpha[G_controlled-1]/np.sum(cost_alpha[G_controlled-1])
PFd=np.sort(PF)
PFd=PFd[0:2]
PFd=-(PFd+0.02)
lambdao = np.zeros((1,1))
rhopo=np.zeros((m,1))
rhomo=np.zeros((m,1))
Eo = np.zeros((n,1))
Eo[:,0]=Vm_orig
x0 = np.concatenate((omegao,thetao,pmecho,loado,Eo),axis=0)

price_signal=np.random.rand(G_controlled.shape[0]+L_controlled.shape[0],1)*80.0-40.0
pricefull=np.zeros((n,1))
price_sup=np.zeros((n,1))
price_sup[np.concatenate((np.transpose(bus_controlled) , L_controlled),axis=0)]=50*np.ones((5,1))
gain=500.0


def md(t,x,n,m,deltaPm,barDeltaP,underDeltaP, M, D, C,Bij,  Bij_nonlinear,Plink_orig, R, Tg, Tb,Td, G, G_controlled,  L,L_controlled, PF,PFd, H, Hg, Hl, rm, Bjj, Tv, xd, xdt, Ef, Bij_vf,CQ,pricefull, bus_controlled,price_sup,gain):
	g = G.shape[0]
	global k_control
	k_control = 0.7*sum(1.0/R[G_controlled])
	k_lambda =1.0/np.mean(np.diag(M[G-1,G-1]))*0.4
	k_rho = 25.0*0.5*np.identity(m)
	omegarange = range(g)
	thetarange = range(g,g+n)
	valverange = range(n+g,n+g+g)
	pmechrange = range(n+1*g,n+1*g+g)
	loadrange = range(n+2*g,n+2*g+n)
	Erange = range(2*n+2*g,2*n+2*g+n)
	omega_g = x[omegarange]
	theta = x[thetarange]
	valve = x[valverange]
	pmech = x[pmechrange]
	load = x[loadrange]
	E=x[Erange]
	edgev=np.ones((m,1));
	for cnt1 in range(m):
		for cnt2 in range(n):
			if C[cnt2,cnt1] !=0:
				edgev[cnt1]=edgev[cnt1]*E[cnt2]
	pgen = np.zeros((n,1)) 
	pmech = pmech.reshape((g,1))
	pgen[G-1] = pmech
	omega = np.zeros((n,1))
	omega[G-1] = omega_g.reshape((g,1))
	DS = D[np.ix_(L-1,L-1)]

	invDS = np.linalg.inv(DS)
	matrixX2 = np.multiply(np.diag(Bij_vf),np.sin(np.matmul(np.transpose(C),theta)))
	matrixX2 = matrixX2.reshape((m,1))
	matrixX3 = np.multiply(edgev,matrixX2)
	Cp = C[np.ix_(L-1,range(m))]
	matrixX = np.matmul(Cp,(matrixX3 -Plink_orig))
	loadL = load[L-1].reshape(l,1)
	loadLC = load[L_controlled-1].reshape(L_controlled.shape[0],1)
	omega[L-1]  =  np.matmul(invDS,( pgen[L-1] + deltaPm[L-1] - loadL - matrixX))
	p_signal = np.zeros((g,1))
	a1 = k_control*PF
	a1 = a1.reshape(bus_controlled.shape[0],1)
#print(price_sup-pricefull)
	a2 = k_control*PFd
	a2 = a2.reshape(L_controlled.shape[0],1)
	p_signal[G_controlled-1] = pricefull[bus_controlled] +  np.divide(gain*np.ones((bus_controlled.shape[0],1)),(price_sup[bus_controlled]-pricefull[bus_controlled]) )  -  np.multiply(np.divide(1.0,(a1)), pmech[G_controlled-1])+ pmech[G_controlled]
	d_signal=np.zeros((n,1))
	d_signal[L_controlled-1]= np.multiply(np.divide(1,(a2)), loadLC) -  pricefull[L_controlled] -  np.divide(gain*np.ones((L_controlled.shape[0],1)),(price_sup[L_controlled]-pricefull[L_controlled])) +loadLC
	loadG = load[G-1]
	loadG = loadG.reshape(g,1)
	a3 = np.matmul(np.diag(Bij_vf.ravel()),np.sin(np.matmul(np.transpose(C),theta)))
	a3 = a3.reshape(m,1)
	b2=np.matmul(C[np.ix_(G-1,range(m))],(np.multiply(edgev,a3) -Plink_orig))
	b3 = np.matmul(D[np.ix_(G-1,G-1)],omega_g)
	b3 = b3.reshape(g,1)
	domega_g  = np.matmul(np.linalg.inv(M[np.ix_(G-1,G-1)]),(deltaPm[G-1] + pgen[G-1] -loadG - b3 - b2));
	dtheta  = omega             
	c1 = -valve.reshape(g,1) +p_signal
	#print(c1.shape)
#dvalve =  np.matmul(np.linalg.inv(np.diag(Tg.ravel())),(c1))
	dpmech  = np.matmul(np.linalg.inv(np.diag(Tb.ravel())),(- pmech + p_signal.reshape(g,1)))
	dload= np.matmul(np.linalg.inv(np.diag(Td.ravel())),(-load.reshape(n,1) + d_signal))
	dlambda = k_lambda*( - np.sum(deltaPm) - np.sum(pgen) + np.sum(load) )
	drhop   = np.matmul(k_rho,(np.matmul(np.transpose(H),(pgen[rm-1]+deltaPm[rm-1]-load[rm-1])) - barDeltaP))
	drhom   = np.matmul(k_rho,(underDeltaP - np.matmul(np.transpose(H),(pgen[rm-1]+deltaPm[rm-1]-load[rm-1]))))
	dE=np.zeros((n,1))
	a4=np.multiply((1-np.multiply((xd-xdt),Bjj[G-1])),E[G-1])
	a4 = a4.reshape(g,1)
	a7=np.matmul(np.diag(Bij_vf.ravel()),np.cos(np.matmul(np.transpose(C),theta)))
	a7 = a7.reshape(m,1)
	#print(edgev.shape)
	a6 = np.multiply(edgev,(a7))
	a5 = np.matmul(CQ[np.ix_(G-1,range(m))],(a6))
	a9 = xd-xdt
	a9 = a9.reshape(g,1)
	EG=E[G-1]
	EG=EG.reshape(g,1)
	a11 = np.divide((a5),EG)
	a8 = np.multiply(a9, a11)
	a12 = np.linalg.inv(np.diag(Tv.ravel()))
	a13 = (Ef*np.ones((g,1))-a4 + a8)
	dE[G-1]=np.matmul(a12,a13)
        #  print(domega_g)
        #print(dtheta)
        #print(dvalve)
        #print(dpmech)
        #print(dload)
        #print(dE)
	y=np.concatenate((domega_g,dtheta,dpmech,dload,dE),axis=0)
	return y
       

t0, t1 = 0, 20                # start and end
t = np.linspace(t0, t1, 100)  # the points of evaluation of solution
x = np.zeros((len(t), len(x0)))   # array for solution
x[0, :] = x0.ravel()
r = integrate.ode(md).set_integrator("lsoda",nsteps=1000,method='bdf')  # choice of method
r.set_initial_value(x0, t0).set_f_params(n,m,deltaPm,barDeltaP,underDeltaP, M, D, C,Bij,  Bij_nonlinear,Plink_orig, R, Tg, Tb,Td, G, G_controlled,  L,L_controlled, PF,PFd, H, Hg, Hl, rm, Bjj, Tv, xd, xdt, Ef, Bij_vf,CQ,pricefull, bus_controlled,price_sup,gain)   # initial values
for i in range(1, t.size):
	xhold =r.integrate(t[i])
	xhold = xhold.ravel()
	x[i, :] = xhold # get one more value, add it to the array
	if not r.successful():
		raise RuntimeError("Could not integrate")

